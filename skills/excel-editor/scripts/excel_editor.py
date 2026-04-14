#!/usr/bin/env python3
"""
Excel Editor Helper Script
Usage: python excel_editor.py <operation_json>

The operation_json is a JSON string describing what to do.
All operations are defined by the caller (Claude) and passed in.

Supported operation types (all via JSON):
  - set_cell_value
  - set_range_values
  - set_cell_format
  - set_range_format
  - set_column_width
  - set_row_height
  - add_sheet
  - rename_sheet
  - delete_sheet
  - reorder_sheets
  - insert_rows
  - insert_columns
  - delete_rows
  - delete_columns
  - apply_filter
  - freeze_panes
  - set_tab_color
  - merge_cells
  - unmerge_cells
  - find_replace
  - set_conditional_format
  - list_workbooks       (discovery)
  - list_sheets          (discovery)
  - read_range           (read data)
"""

import sys
import json
import os
import glob
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
        numbers as num_styles
    )
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule, FormulaRule
    from openpyxl.styles.differential import DifferentialStyle
except ImportError:
    print(json.dumps({"error": "openpyxl not installed. Run: pip install openpyxl"}))
    sys.exit(1)


def parse_color(color_str):
    """Convert color name or hex to openpyxl ARGB hex string."""
    color_map = {
        "red": "FFFF0000", "green": "FF00FF00", "blue": "FF0000FF",
        "yellow": "FFFFFF00", "orange": "FFFF8000", "purple": "FF800080",
        "black": "FF000000", "white": "FFFFFFFF", "gray": "FF808080",
        "grey": "FF808080", "lightblue": "FFADD8E6", "lightgreen": "FF90EE90",
        "lightgray": "FFD3D3D3", "lightgrey": "FFD3D3D3", "cyan": "FF00FFFF",
        "magenta": "FFFF00FF", "pink": "FFFFC0CB", "darkblue": "FF00008B",
        "darkgreen": "FF006400", "darkred": "FF8B0000", "gold": "FFFFD700",
        "teal": "FF008080", "navy": "FF000080", "lime": "FF00FF00",
        "brown": "FFA52A2A", "silver": "FFC0C0C0",
    }
    if color_str is None:
        return None
    s = str(color_str).strip().lower()
    if s in color_map:
        return color_map[s]
    s = s.lstrip("#")
    if len(s) == 6:
        return "FF" + s.upper()
    if len(s) == 8:
        return s.upper()
    return s.upper()


def get_cell(ws, ref):
    """Get a cell by A1 or row/col reference."""
    if isinstance(ref, str):
        return ws[ref]
    return ws.cell(row=ref[0], column=ref[1])


def apply_style(cell, style_dict):
    """Apply style properties to a cell."""
    if not style_dict:
        return

    if "font" in style_dict:
        f = style_dict["font"]
        kwargs = {}
        if "bold" in f: kwargs["bold"] = f["bold"]
        if "italic" in f: kwargs["italic"] = f["italic"]
        if "underline" in f: kwargs["underline"] = f.get("underline", "single") if f.get("underline") else None
        if "size" in f: kwargs["size"] = f["size"]
        if "color" in f: kwargs["color"] = parse_color(f["color"])
        if "name" in f: kwargs["name"] = f["name"]
        if "strikethrough" in f: kwargs["strike"] = f["strikethrough"]
        existing = cell.font
        cell.font = Font(
            bold=kwargs.get("bold", existing.bold),
            italic=kwargs.get("italic", existing.italic),
            underline=kwargs.get("underline", existing.underline),
            size=kwargs.get("size", existing.size),
            color=kwargs.get("color", existing.color.rgb if existing.color else None),
            name=kwargs.get("name", existing.name),
            strike=kwargs.get("strike", existing.strike),
        )

    if "fill" in style_dict:
        fill = style_dict["fill"]
        color = parse_color(fill.get("color") or fill.get("fgColor"))
        if color:
            cell.fill = PatternFill(
                fill_type=fill.get("type", "solid"),
                fgColor=color,
            )
        elif fill.get("clear"):
            cell.fill = PatternFill(fill_type=None)

    if "alignment" in style_dict:
        a = style_dict["alignment"]
        kwargs = {k: v for k, v in {
            "horizontal": a.get("horizontal"),
            "vertical": a.get("vertical"),
            "wrap_text": a.get("wrap_text"),
            "shrink_to_fit": a.get("shrink_to_fit"),
            "text_rotation": a.get("text_rotation"),
            "indent": a.get("indent"),
        }.items() if v is not None}
        cell.alignment = Alignment(**kwargs)

    if "number_format" in style_dict:
        cell.number_format = style_dict["number_format"]

    if "border" in style_dict:
        b = style_dict["border"]
        def make_side(spec):
            if spec is None:
                return Side()
            if isinstance(spec, str):
                return Side(border_style=spec)
            return Side(
                border_style=spec.get("style", "thin"),
                color=parse_color(spec.get("color", "FF000000")),
            )
        cell.border = Border(
            left=make_side(b.get("left")),
            right=make_side(b.get("right")),
            top=make_side(b.get("top")),
            bottom=make_side(b.get("bottom")),
            outline=b.get("outline", False),
        )


def resolve_sheet(wb, sheet_ref):
    """Get worksheet by name or index (0-based)."""
    if sheet_ref is None:
        return wb.active
    if isinstance(sheet_ref, int):
        return wb.worksheets[sheet_ref]
    return wb[sheet_ref]


def find_workbooks(folder, pattern="*.xlsx"):
    """Find Excel files in a folder."""
    folder = Path(folder).expanduser()
    results = []
    for ext in ("*.xlsx", "*.xlsm", "*.xls"):
        results.extend(folder.glob(ext))
    return sorted(set(results))


def run_operation(op):
    """Execute a single operation dict. Returns result dict."""
    op_type = op.get("type")
    folder = op.get("folder")
    workbook = op.get("workbook")  # filename or None for all
    sheet = op.get("sheet")        # sheet name, index, or None for active

    # Discovery operations (no file write needed)
    if op_type == "list_workbooks":
        files = find_workbooks(folder)
        return {"workbooks": [f.name for f in files]}

    if op_type == "list_sheets":
        wb_path = Path(folder).expanduser() / workbook
        wb = openpyxl.load_workbook(wb_path, data_only=False)
        info = []
        for ws in wb.worksheets:
            info.append({"name": ws.title, "max_row": ws.max_row, "max_column": ws.max_column})
        return {"sheets": info}

    if op_type == "read_range":
        wb_path = Path(folder).expanduser() / workbook
        wb = openpyxl.load_workbook(wb_path, data_only=True)
        ws = resolve_sheet(wb, sheet)
        cell_range = op.get("range")
        if cell_range:
            data = []
            for row in ws[cell_range]:
                data.append([c.value for c in row])
        else:
            data = [[c.value for c in row] for row in ws.iter_rows()]
        return {"sheet": ws.title, "data": data}

    # Write operations
    target_files = []
    if workbook:
        target_files = [Path(folder).expanduser() / workbook]
    else:
        target_files = find_workbooks(folder)

    results = []
    for wb_path in target_files:
        wb = openpyxl.load_workbook(wb_path)
        changed = False

        # Determine target sheets
        if sheet == "__all__" or (sheet is None and op_type in ("find_replace",)):
            target_sheets = wb.worksheets
        elif sheet is not None:
            target_sheets = [resolve_sheet(wb, sheet)]
        else:
            target_sheets = [wb.active]

        for ws in target_sheets:
            if op_type == "set_cell_value":
                cell = get_cell(ws, op["cell"])
                cell.value = op["value"]
                changed = True

            elif op_type == "set_range_values":
                start_row = op.get("start_row", 1)
                start_col = op.get("start_col", 1)
                if isinstance(start_col, str):
                    start_col = column_index_from_string(start_col)
                for r_idx, row_data in enumerate(op["values"]):
                    for c_idx, val in enumerate(row_data):
                        ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=val)
                changed = True

            elif op_type == "set_cell_format":
                for cell_ref in (op.get("cells") or [op.get("cell")]):
                    apply_style(ws[cell_ref], op.get("style", {}))
                changed = True

            elif op_type == "set_range_format":
                for row in ws[op["range"]]:
                    for cell in row:
                        apply_style(cell, op.get("style", {}))
                changed = True

            elif op_type == "set_column_width":
                cols = op.get("columns") or [op.get("column")]
                width = op.get("width")
                for col in cols:
                    ws.column_dimensions[col].width = width
                changed = True

            elif op_type == "set_row_height":
                rows = op.get("rows") or [op.get("row")]
                height = op.get("height")
                for row in rows:
                    ws.row_dimensions[row].height = height
                changed = True

            elif op_type == "freeze_panes":
                ws.freeze_panes = op.get("cell", "A1")
                changed = True

            elif op_type == "merge_cells":
                ws.merge_cells(op["range"])
                changed = True

            elif op_type == "unmerge_cells":
                ws.unmerge_cells(op["range"])
                changed = True

            elif op_type == "insert_rows":
                ws.insert_rows(op["row"], op.get("count", 1))
                changed = True

            elif op_type == "insert_columns":
                ws.insert_cols(op["col"] if isinstance(op["col"], int)
                               else column_index_from_string(op["col"]),
                               op.get("count", 1))
                changed = True

            elif op_type == "delete_rows":
                ws.delete_rows(op["row"], op.get("count", 1))
                changed = True

            elif op_type == "delete_columns":
                ws.delete_cols(op["col"] if isinstance(op["col"], int)
                               else column_index_from_string(op["col"]),
                               op.get("count", 1))
                changed = True

            elif op_type == "apply_filter":
                ws.auto_filter.ref = op.get("range", ws.dimensions)
                changed = True

            elif op_type == "find_replace":
                import re
                find_val = op["find"]
                replace_val = op["replace"]
                match_case = op.get("match_case", False)
                flags = 0 if match_case else re.IGNORECASE
                count = 0
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is None:
                            continue
                        cell_str = str(cell.value)
                        new_str = re.sub(re.escape(find_val), replace_val, cell_str, flags=flags)
                        if new_str != cell_str:
                            cell.value = new_str
                            count += 1
                changed = changed or count > 0

            elif op_type == "set_tab_color":
                ws.sheet_properties.tabColor = parse_color(op["color"]).lstrip("FF") if parse_color(op["color"]) else None
                changed = True

        # Sheet-level operations on the workbook
        if op_type == "add_sheet":
            name = op["name"]
            pos = op.get("position")
            if name not in wb.sheetnames:
                if pos is not None:
                    wb.create_sheet(name, pos)
                else:
                    wb.create_sheet(name)
                changed = True

        elif op_type == "rename_sheet":
            ws = resolve_sheet(wb, op["sheet"])
            ws.title = op["new_name"]
            changed = True

        elif op_type == "delete_sheet":
            if op["sheet"] in wb.sheetnames:
                del wb[op["sheet"]]
                changed = True

        elif op_type == "reorder_sheets":
            order = op["order"]  # list of sheet names in desired order
            wb._sheets = [wb[name] for name in order if name in wb.sheetnames]
            changed = True

        if changed:
            wb.save(wb_path)
            results.append({"file": wb_path.name, "status": "saved"})
        else:
            results.append({"file": wb_path.name, "status": "no_change"})

    return {"results": results}


def main():
    if len(sys.argv) < 2:
        print(json.dumps({"error": "Usage: python excel_editor.py '<operation_json>'"}))
        sys.exit(1)

    try:
        op = json.loads(sys.argv[1])
    except json.JSONDecodeError as e:
        print(json.dumps({"error": f"Invalid JSON: {e}"}))
        sys.exit(1)

    try:
        result = run_operation(op)
        print(json.dumps(result, default=str))
    except Exception as e:
        import traceback
        print(json.dumps({"error": str(e), "traceback": traceback.format_exc()}))
        sys.exit(1)


if __name__ == "__main__":
    main()
